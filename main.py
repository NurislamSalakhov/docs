from docxtpl import DocxTemplate
import openpyxl

# Give the location of the file
path = "book4.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active


name_org = sheet_obj.cell(row=1, column=2).value
name_service = sheet_obj.cell(row=2, column=2).value
amount_people = sheet_obj.cell(row=3, column=2).value
date = sheet_obj.cell(row=4, column=2).value
place_with_time = sheet_obj.cell(row=5, column=2).value
amount_money = sheet_obj.cell(row=6, column=2).value
data = sheet_obj.cell(row=7, column=2).value
name_signer = sheet_obj.cell(row=8, column=2).value
time_s = sheet_obj.cell(row=9, column=2).value

doc = DocxTemplate("file.docx")

context = {
    'company': name_org,
    'dictionary': name_service,
    'quantity': amount_people,
    'time': date,
    'place': place_with_time,
    'money': amount_money,
    'hz': data,
    'name': name_signer,
    'time_s': time_s,
    'money_divided': amount_money//amount_people
}

doc.render(context)

doc.save(f"{name_org[:9]}.docx")
